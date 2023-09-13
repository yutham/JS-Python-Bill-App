from flask import Flask, render_template
import openpyxl  # This Import is for using excel is python
import pandas as pd  # Its For creating data frames
import webview

import sys
import os


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


# wb_path = "./Product.xlsx"
# os.chdir("E:/Bill App/JS-Python-Bill-App-main")
wb = openpyxl.load_workbook(resource_path("Product.xlsx"))
# def load_workbook(wb_path):  # Loading Excel Sheet
#     if os.path.exists(wb_path):
#         return openpyxl.load_workbook(wb_path)
#     else:
#         return None
# Get the current directory of the script
# xlsx_filename = 'Product.xlsx'

# current_directory = os.path.dirname(os.path.abspath(xlsx_filename))
# print(current_directory)

# Construct the full path to the Excel file
# wb_path = "/Product.xlsx"
# print(wb_path)


# wb = openpyxl.load_workbook(wb_path)
# print(wb)


# wb = openpyxl.load_workbook(wb_path)
sheet = wb['Sheet1']
sheet_obj = wb.active

# Print value of cell object
# using the value attribute
# print(list(sheet.values))
# print(list(sheet.values))
# print(sheet_obj.cell)


def Add_Data():
    try:
        df = pd.read_excel("./Product.xlsx")
        max_value = df['Id'].max()
        print(type(max_value))
        if not isinstance(max_value, int):
            max_value = 1
        else:
            max_value += 1
        new_record = {'Id': max_value,
                      'Column1': 'Value1',
                      'Column2': 'Value2'}
        new_record['Column1'] = input("\nEnter the Product Name: ")
        new_record['Column2'] = int(input("Enter the Rate: "))
        sheet_obj.append([new_record['Id'],
                          new_record['Column1'],
                          new_record['Column2']])
        wb.save(wb_path)
        print("Added Successfully")
        add_more = input("\nWant to Add more Data? (yes/no): ")
        if add_more.lower() == "yes":
            Add_Data()
    except Exception as e:
        print(f"An error occurred while adding data: {str(e)}")


def view():
    # viewing the data using data-frame
    return list(sheet.values)


def search(id):  # this function is to search the data in the excel sheet
    max_row = sheet.max_row
    for i in range(1, max_row+1):
        if sheet.cell(row=i, column=1).value == id:
            print("record Found")
            return i


def Display_record(row):  # this function is to the single data in the excel sheet
    try:
        max_column = sheet.max_column
        for i in range(1, max_column+1):
            cell_obj = sheet.cell(row=row, column=i)
            print(cell_obj.value)
    except Exception as e:
        print(f"An error occurred while displaying the record: {str(e)}")


def Update_records(row):  # this function is to update the data in the excel sheet
    try:
        new_record = input("\n Enter the Product Name:")
        new_rate = int(input("\n Enter the Rate:"))
    except ValueError:
        print("Invalid Rate. Rate should be an integer.")
        return
    sheet.cell(row=row, column=2, value=new_record)
    sheet.cell(row=row, column=3, value=new_rate)
    wb.save(wb_path)
    return print("record updated")


def Delete_record(row):  # this function is to delete the data in the excel sheet
    try:
        sheet.cell(row=row, column=2, value="")
        sheet.cell(row=row, column=3, value="")
        print("Record Deleted Successfully")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    wb.save(wb_path)
    return


# import webview # Libray to create a gui window


app = Flask(__name__, static_folder='./static', template_folder='./templates')


@app.route('/')
def login():
    return render_template("login.html", name='cairocoders Home page')


@app.route('/home')
def home():
    return render_template("home.html", name='cairocoders page 2', message=view())


@app.route('/Admin', methods=["POST"])
def viewinpage():
    return 'hello'


# @app.route('/Admin')
# def admin():
#     while True:
#         print("\n 1:add")
#         print("\n 2:view")
#         print("\n 3:Update")
#         print("\n 4:Delete")
#         ch = int(input("enter your choic:"))
#         if ch == 1:
#             Add_Data()
#         elif ch == 2:
#             view()
#         elif ch == 3:
#             x = int(input("\n Enter the Id:"))
#             row = search(x)
#             Display_record(row)
#             y = input("\n want to edit the record ? yes/no:")
#             if y.lower() == 'yes':
#                 Update_records(row)
#         elif ch == 4:
#             x = int(input("\n Enter the Id:"))
#             row = search(x)
#             Display_record(row)
#             y = input("\n want to Delete the record ? yes/no:")
#             if y.lower() == 'yes':
#                 Delete_record(row)


webview.create_window('Billing App', app)  # To enable a gui window


if __name__ == '__main__':
    # app.run(debug=True)  # Only use for development (recommended)
    webview.start()  # To start a gui
