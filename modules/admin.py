import openpyxl  # This Import is for using excel is python
import os
import sys
import shutil


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def get_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        if hasattr(sys, "_MEIPASS"):
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            abs_home = os.path.abspath(os.path.expanduser("~"))
            abs_dir_app = os.path.join(abs_home, f".bill_app")
            if not os.path.exists(abs_dir_app):
                os.mkdir(abs_dir_app)
                shutil.copy(resource_path(relative_path),
                            os.path.join(abs_dir_app, relative_path))
            excel_path = os.path.join(abs_dir_app, relative_path)
        else:
            # If running from the script's directory
            excel_path = os.path.abspath(relative_path)
    except Exception:
        excel_path = os.path.abspath(".")

    return excel_path


excel_path = get_path("stocks.xlsx")
wb = openpyxl.load_workbook(excel_path)


sheet = wb['Sheet1']
sheet_obj = wb.active


def Add_Stock(New_Record):  # To Add Stock

    try:
        index = int(sheet_obj.max_row)-1
        print(sheet_obj)
        if not isinstance(index, int):
            index = 1
        else:
            index += 1

        sheet_obj.append([index,
                          New_Record['Product Name'],
                          int(New_Record['Price']),
                          int(New_Record['Quantity'])])
        wb.save(excel_path)
        # print("Data Added")

        return True

    except Exception as e:
        print(f"An error occurred while adding data: {str(e)}")
        return False


def Display_stocks():  # To Display Record
    try:
        data = []
        header = [cell.value for cell in sheet_obj[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(header, row))
            data.append(row_data)

        # print(data)
        return data
    except Exception as e:
        print(f"An error occurred while adding data: {str(e)}")
        return False


def Edit_stocks(index, Edit_Record):  # this function is to update the data in the excel sheet
    try:
        row = index+1
        sheet.cell(row, column=2, value=Edit_Record['Product Name'])
        sheet.cell(row, column=3, value=int(Edit_Record['Price']))
        sheet.cell(row, column=4, value=int(Edit_Record['Quantity']))
        wb.save(excel_path)
        return True
    except Exception as e:
        print(f"An error occurred while adding data: {str(e)}")
        return False


def Delete_stocks(index):  # this function is to delete the data in the excel sheet
    try:
        row = index+1
        sheet_obj.delete_rows(row)
        max_index = sheet_obj.max_row
        for i in range(index, max_index):
            sheet.cell(row, column=1, value=i)
            row += 1

        wb.save(excel_path)
        return True
    except Exception as e:
        print(f"An error occurred while adding data: {str(e)}")
        return False
